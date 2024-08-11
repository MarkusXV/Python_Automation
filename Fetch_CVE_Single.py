import requests
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def fetch_cve(api_url, cve):
  try:
    print(f"{api_url}?cveId={cve}")
    response = requests.get(f"{api_url}?cveId={cve}")
    response.raise_for_status()  # Raise an exception for HTTP errors
    data = response.json()
    if 'vulnerabilities' in data:
      return data['vulnerabilities']
    else:
      print("No CVE data found in the response.")
      return []
  except requests.RequestException as e:
    print(f"HTTP request failed: {e}")
    return []

def update_cve_info(api_url, cve, filename):
  cve_items = fetch_cve(api_url, cve)
  if not cve_items: 
    print("No CVEs found")
    return

  wb = Workbook()
  ws = wb.active
  ws.title = f"CVE Info {cve}"

  headers = ["CVE ID", "Description", "CVSS"]
  ws.append(headers)


  ws.column_dimensions["A"].width = 14
  ws.column_dimensions["B"].width = 10
  ws.column_dimensions["C"].width = 200

  for col in ws.iter_cols(min_row=2, min_col=3):
    for cell in col:
      cell.alignment = Alignment(wrap_text=True)

  row_num = 2
  for item in cve_items:
    cve_id = item['cve']['id']
    description = item['cve']['descriptions'][0]['value'] if item['cve']['descriptions'] else "N/A"
    try:
      cvss_score = item['cve']['metrics']['cvssMetricV31'][0]['cvssData']['baseScore']
    except:
      try:
        cvss_score = item['cve']['metrics']['cvssMetricV30'][0]['cvssData']['baseScore']
      except:
          try:
            cvss_score = item['cve']['metrics']['cvssMetricV2'][0]['cvssData']['baseScore']
          except:
            print("No V2")
            cvss_score = "N/A"
    row = [cve_id, description, cvss_score]
    ws.append([cve_id, cvss_score, description])
    ws.row_dimensions[row_num].height = 32
    row_num += 1

  wb.save(filename)
  print(f"CVE information updated for the specified CVE in {filename}")

# Example usage:
api_url = 'https://services.nvd.nist.gov/rest/json/cves/2.0'
cve = input('What CVE do you want?\n').upper()
filename = 'cve_info.xlsx'
update_cve_info(api_url, cve, filename)
